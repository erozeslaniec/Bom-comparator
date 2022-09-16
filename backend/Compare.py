from PySide6.QtCore import QObject, Slot, Signal
import openpyxl
import numpy as np
import pyperclip

class Compare(QObject):
    firstFile=''
    secondFile=''

    def __init__(self):
        self.firstFile=''
        self.secondFile=''
        super().__init__()

    submitText1Signal = Signal(str)
    submitText2Signal = Signal(str)







    @Slot(result=list)
    def compareReq(self):
        def createString(ls):
            tStr = ''
            lp = 1
            for item in ls:
                tempLineStr = "{}\t\"{}\"\t\"{}\"\t\"{}\"\t\"{}\"\n".format(str(lp), str(item[0]), str(item[1]),
                                                                            str(item[2]), str(item[3]))
                tempLineStr = tempLineStr.replace("None", "")
                tStr = tStr + tempLineStr
                lp += 1
            return tStr

        def addItemsToList(filePath):
            filledList = []
            wb_obj = openpyxl.load_workbook(filePath)
            sheet_obj = wb_obj.active
            row = sheet_obj.max_row
            column = sheet_obj.max_column

            for i in range(row):
                if (i == 0):
                    continue
                else:
                    listOfElements = []
                    for j in range(column):
                        if (j == 0):
                            continue
                        else:
                            listOfElements.append((sheet_obj.cell(row=i + 1, column=j + 1)).value)
                    filledList.append(listOfElements)
            return filledList

        def createDifferenceList(prevItems, currItems):
            diffItems = []
            for i in currItems:
                if i[0] in prevItems:  # in (x[0] for x in previousItems):
                    index = np.where(prevItems == i[0])[0]
                    list = [prevItems[index[0]][0], prevItems[index[0]][1], prevItems[index[0]][2],
                            int(i[3]) - int(prevItems[index[0]][3]), prevItems[index[0]][4]]
                    diffItems.append(list)
                else:
                    diffItems.append(i)
            for i in prevItems:
                if not (i[0] in currItems):
                    list = [i[0], i[1], i[2], -1 * int(i[3]), i[4]]
                    diffItems.append(list)
            return diffItems

        def putListInClipboard(clipList):
            differenceItemsText = createString(clipList)
            pyperclip.copy(differenceItemsText)



        previousItems=[]
        currentItems=[]
        differenceItems=[]

        previousItems=np.array(addItemsToList(self.firstFile))
        currentItems=np.array(addItemsToList(self.secondFile))

        differenceItems=createDifferenceList(previousItems, currentItems)
        print("Success")
        putListInClipboard(differenceItems)

    @Slot(str)
    def addFirstFile(self, surl):
        surl=str(surl).replace("file:///", "")
        self.firstFile = surl
        self.submitText1Signal.emit(surl)
        print(surl + " added")

    @Slot(str)
    def addSecondFile(self, surl):
        surl = str(surl).replace("file:///", "")
        self.secondFile = surl
        self.submitText2Signal.emit(surl)
        print(surl + " added")


